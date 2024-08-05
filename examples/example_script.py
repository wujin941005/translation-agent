import os
import re
import translation_agent as ta
from openpyxl import Workbook, load_workbook
from datetime import date
from icecream import ic

def count_words(text):
    # 去除首尾空白字符
    text = text.strip()
    
    # 专门方法：统计英文单词、数字和中文字符
    english_words = len(re.findall(r'\b[a-zA-Z0-9]+\b', text))
    chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', text))
    specialized_count = english_words + chinese_chars
    
    # 通用方法：统计所有非空白字符序列
    general_count = len(re.findall(r'\S+', text))
    
    # 返回两种方法中的最大值
    return max(specialized_count, general_count)

def create_or_update_excel(client_name, source_text, source_lang, target_lang, translation):
    filename = f"{client_name}.xlsx"
    
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = ["日期", "源语言", "目标语言", "源文字数", "原文", "译文"]
        ws.append(headers)
    
    # 添加新行
    new_row = [
        date.today().strftime("%Y-%m-%d"),
        source_lang,
        target_lang,
        count_words(source_text),
        source_text,
        translation
    ]
    ws.append(new_row)
    
    wb.save(filename)

if __name__ == "__main__":
    source_lang, target_lang, country = "English", "Chinese", "China"
    client_name = "客户名称"  # 这里可以手动设置客户名称

    relative_path = "sample-texts/sample-short1.txt"
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    full_path = os.path.join(script_dir, relative_path)
    
    with open(full_path, encoding="utf-8") as file:
        source_text = file.read()
    
    print(f"Source text:\n\n{source_text}\n------------\n")
    
    translation = ta.translate(
        source_lang=source_lang,
        target_lang=target_lang,
        source_text=source_text,
        country=country,
        api_choice="claude",
        user_prompt=""
    )
    
    print(f"Translation:\n\n{translation}")

    # 创建或更新Excel文件
    create_or_update_excel(client_name, source_text, source_lang, target_lang, translation)
    ic("Excel file has been updated.")